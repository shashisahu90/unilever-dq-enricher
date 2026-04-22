import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Unilever DQ Enricher", page_icon="📦", layout="wide")
st.title("📦 Unilever – Weekly Raw File Enricher")
st.markdown("Upload the weekly **Data Quality by Carrier** raw file to generate the enriched Excel report.")

# ── Colours ────────────────────────────────────────────────────
DARK_BLUE = "1F3864"
MID_BLUE  = "2E75B6"
LT_BLUE   = "D6E4F0"
GREEN     = "C6EFCE"
YELLOW    = "FFEB9C"
RED       = "FFC7CE"
WHITE     = "FFFFFF"
GREY      = "F2F2F2"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10): return Font(name="Arial", bold=bold, color=color, size=size)
def _align(h="left", wrap=False): return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
def _bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, txt, bg=DARK_BLUE, fg=WHITE, bold=True, size=10, align="center", wrap=False):
    cell.value = txt; cell.font = _font(bold, fg, size)
    cell.fill = _fill(bg); cell.alignment = _align(align, wrap); cell.border = _bdr()

def val(cell, v, bg=WHITE, bold=False, size=10, align="left", color="000000", wrap=False):
    cell.value = v; cell.font = _font(bold, color, size)
    cell.fill = _fill(bg); cell.alignment = _align(align, wrap); cell.border = _bdr()

# ── Enrichment ─────────────────────────────────────────────────
M1_COL = "Pickup Arrival Milestone (UTC)"
M2_COL = "Pickup Departure Milestone (UTC)"
M3_COL = "Final Destination Arrival Milestone (UTC)"
M4_COL = "Final Destination Departure Milestone (UTC)"

def _has_milestone(v):
    """Return True if the cell has a real timestamp value (not empty/nan/0/NaT)."""
    if v is None: return False
    s = str(v).strip().lower()
    return s not in ("", "nan", "none", "0", "nat", "0.0")

def enrich(df_raw):
    df = df_raw.copy()
    df.columns = [c.strip() for c in df.columns]

    # Ensure milestone columns exist (avoid KeyError on edge cases)
    for col in [M1_COL, M2_COL, M3_COL, M4_COL]:
        if col not in df.columns:
            df[col] = ""

    # ── FIX 1: Milestones Completeness ─────────────────────────
    # Count exactly how many of the 4 milestone columns have a real timestamp value
    # Result: "0/4", "1/4", "2/4", "3/4", "4/4" only
    df["Milestones Completeness"] = df.apply(
        lambda row: "{}/4".format(
            sum([
                _has_milestone(row[M1_COL]),
                _has_milestone(row[M2_COL]),
                _has_milestone(row[M3_COL]),
                _has_milestone(row[M4_COL]),
            ])
        ), axis=1
    )

    # ── FIX 2: Milestones Reached ───────────────────────────────
    # Based on Milestones Completeness, check WHICH of the 4 columns have a value:
    #   M1 = Pickup Arrival, M2 = Pickup Departure,
    #   M3 = Final Destination Arrival, M4 = Final Destination Departure
    # If 0/4 → "None"
    MILESTONE_MAP = [
        (M1_COL, "M1"),
        (M2_COL, "M2"),
        (M3_COL, "M3"),
        (M4_COL, "M4"),
    ]
    def milestone_reached(row):
        mc = str(row.get("Milestones Completeness", "0/4")).strip()
        if mc == "0/4":
            return "None"
        reached = [label for col, label in MILESTONE_MAP if _has_milestone(row[col])]
        return ",".join(reached) if reached else "None"

    df["Milestones Reached"] = df.apply(milestone_reached, axis=1)

    # ── FIX 3: Tracking Status ──────────────────────────────────
    # Tracked = TRUE  → 4/4 = "Fully Tracked", anything else = "Partially Tracked"
    # Tracked = FALSE → RCA from "Tracking Error" column (already populated in raw file)
    df["Tracked_bool"] = df["Tracked"].astype(str).str.upper().str.strip() == "TRUE"

    def tracking_status(row):
        if row["Tracked_bool"]:
            # TRUE: determine by milestone completeness
            mc = str(row.get("Milestones Completeness", "0/4")).strip()
            return "Fully Tracked" if mc == "4/4" else "Partially Tracked"
        else:
            # FALSE: use Tracking Error as the RCA label
            te = str(row.get("Tracking Error", "")).strip()
            if te and te.lower() not in ("nan", "none", ""):
                return te
            # fallback: Final Status Reason (skip generic TIMED_OUT)
            fsr = str(row.get("Final Status Reason", "")).strip()
            if fsr and fsr.lower() not in ("nan", "none", "", "timed_out"):
                return fsr
            return "Not Tracked"

    df["Tracking Status"] = df.apply(tracking_status, axis=1)

    if "Carrier Comments" not in df.columns:
        df["Carrier Comments"] = ""
    return df

# ── Sheet 1: Pivot for Tracker ─────────────────────────────────
def sheet_pivot(ws, df):
    ws.title = "Pivot for Tracker"
    ws.sheet_view.showGridLines = False

    pt = (df.groupby(["Carrier Name","Tracked_bool"]).size()
            .unstack(fill_value=0)
            .rename(columns={False:"FALSE",True:"TRUE"}))
    for c in ["FALSE","TRUE"]:
        if c not in pt.columns: pt[c] = 0
    pt["Grand Total"] = pt["FALSE"] + pt["TRUE"]
    pt = pt.sort_index()

    all_st = sorted(df["Tracking Status"].dropna().unique().tolist())
    ps = (df.groupby(["Carrier Name","Tracking Status"]).size()
            .unstack(fill_value=0))
    for s in all_st:
        if s not in ps.columns: ps[s] = 0
    ps = ps[all_st]
    ps["Grand Total"] = ps.sum(axis=1)
    ps = ps.sort_index()

    ws.column_dimensions["A"].width = 40
    for i in range(3): ws.column_dimensions[get_column_letter(2+i)].width = 13
    GAP = 6; RIGHT = GAP + 1
    ws.column_dimensions[get_column_letter(GAP)].width = 3
    ws.column_dimensions[get_column_letter(RIGHT)].width = 40
    for i in range(len(all_st)+1):
        ws.column_dimensions[get_column_letter(RIGHT+1+i)].width = 20

    ws.merge_cells(start_row=1,end_row=1,start_column=1,end_column=4)
    hdr(ws.cell(1,1),"Count of Tracked",DARK_BLUE,WHITE,True,11)
    ws.merge_cells(start_row=1,end_row=1,start_column=RIGHT,end_column=RIGHT+len(all_st))
    hdr(ws.cell(1,RIGHT),"Count of Tracking Status",DARK_BLUE,WHITE,True,11)
    ws.row_dimensions[1].height = 25

    for i,t in enumerate(["Row Labels","FALSE","TRUE","Grand Total"],1):
        hdr(ws.cell(2,i),t,MID_BLUE,WHITE,True,10)
    for i,t in enumerate(["Row Labels"]+all_st+["Grand Total"]):
        hdr(ws.cell(2,RIGHT+i),t,MID_BLUE,WHITE,True,10,wrap=True)
    ws.row_dimensions[2].height = 35

    carriers = pt.index.tolist()
    for r,carrier in enumerate(carriers):
        row = r+3; bg = GREY if r%2==0 else WHITE
        rt = pt.loc[carrier]
        val(ws.cell(row,1),carrier,bg,True)
        val(ws.cell(row,2),int(rt.get("FALSE",0)),bg,align="center")
        val(ws.cell(row,3),int(rt.get("TRUE",0)),bg,align="center")
        val(ws.cell(row,4),int(rt["Grand Total"]),bg,True,align="center")
        rs = ps.loc[carrier] if carrier in ps.index else pd.Series(dtype=int)
        val(ws.cell(row,RIGHT),carrier,bg,True)
        for i,s in enumerate(all_st):
            v = int(rs.get(s,0))
            val(ws.cell(row,RIGHT+1+i),v if v else "",bg,align="center")
        val(ws.cell(row,RIGHT+1+len(all_st)),int(rs.sum()) if len(rs) else 0,bg,True,align="center")

    gr = len(carriers)+3
    val(ws.cell(gr,1),"Grand Total",LT_BLUE,True)
    val(ws.cell(gr,2),int(pt["FALSE"].sum()),LT_BLUE,True,align="center")
    val(ws.cell(gr,3),int(pt["TRUE"].sum()),LT_BLUE,True,align="center")
    val(ws.cell(gr,4),int(pt["Grand Total"].sum()),LT_BLUE,True,align="center")
    val(ws.cell(gr,RIGHT),"Grand Total",LT_BLUE,True)
    for i,s in enumerate(all_st):
        val(ws.cell(gr,RIGHT+1+i),int(ps[s].sum()),LT_BLUE,True,align="center")
    val(ws.cell(gr,RIGHT+1+len(all_st)),int(ps["Grand Total"].sum()),LT_BLUE,True,align="center")

# ── Sheet 2: Summary ───────────────────────────────────────────
def sheet_summary(ws, df):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    total         = len(df)
    not_tracked   = len(df[df["Tracked_bool"]==False])
    tracked_count = len(df[df["Tracked_bool"]==True])
    fully         = len(df[df["Tracking Status"]=="Fully Tracked"])
    partially     = len(df[df["Tracking Status"]=="Partially Tracked"])
    rca_counts    = df[df["Tracked_bool"]==False]["Tracking Status"].value_counts().to_dict()

    for i,t in enumerate(["P44 Analysis","Count","% of Total"],1):
        hdr(ws.cell(2,i),t,MID_BLUE,WHITE,True,10)
    ws.row_dimensions[2].height = 20

    rows = [("FALSE (Not Tracked)",not_tracked,"FFC7CE",True)]
    for rca,cnt in sorted(rca_counts.items(),key=lambda x:-x[1]):
        rows.append(("  "+rca,cnt,"FCE4D6",False))
    rows.append(("TRUE (Tracked)",tracked_count,"C6EFCE",True))
    rows.append(("  Fully Tracked",fully,"E2EFDA",False))
    rows.append(("  Partially Tracked",partially,"FFEB9C",False))
    rows.append(("Grand Total",total,"D6E4F0",True))

    for r,(label,count,bg,bold) in enumerate(rows):
        row = r+3
        pct = f"{count/total*100:.1f}%" if total>0 else "0.0%"
        val(ws.cell(row,1),label,bg,bold)
        val(ws.cell(row,2),count,bg,bold,align="center")
        val(ws.cell(row,3),pct,bg,bold,align="center")
        ws.row_dimensions[row].height = 18

# ── Sheet 3: FTL data Quality ──────────────────────────────────
def sheet_ftl(ws, df):
    ws.title = "FTL data Quality"
    ws.sheet_view.showGridLines = False

    COLS = [
        ("Shipment Created (UTC)",22),("Customer Tenant Name",16),
        ("Carrier Name",30),("Bill of Lading",16),("Order Number",16),
        ("Pickup Appointement Window (UTC)",28),("Delivery Appointement Window (UTC)",28),
        ("Tracked",10),("Connection Type",16),("Tracking Method",22),
        ("Active Equipment ID",22),("Historical Equipment ID",22),
        ("Pickup City State",22),("Final Destination City State",22),
        ("Tracking Window Start (UTC)",22),("Tracking Window End (UTC)",22),
        ("Pickup Arrival Milestone (UTC)",22),("Pickup Departure Milestone (UTC)",22),
        ("Final Destination Arrival Milestone (UTC)",22),
        ("Final Destination Departure Milestone (UTC)",22),
        ("Milestones Completeness",18),("Milestones Reached",18),
        ("Tracking Status",30),
        ("Milestone Error 1",28),("Milestone Error 2",28),("Milestone Error 3",28),
        ("Carrier Comments",30),
    ]
    STATUS_BG = {"Fully Tracked":"C6EFCE","Partially Tracked":"FFEB9C"}

    for ci,(cname,width) in enumerate(COLS,1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        hdr(ws.cell(1,ci),cname,DARK_BLUE,WHITE,True,10,wrap=True)
    ws.row_dimensions[1].height = 30

    for r,(_, row_data) in enumerate(df.iterrows()):
        er = r+2
        status  = str(row_data.get("Tracking Status",""))
        tracked = str(row_data.get("Tracked","")).upper().strip()
        row_bg  = STATUS_BG.get(status, "FCE4D6") if tracked!="TRUE" else STATUS_BG.get(status,WHITE)
        for ci,(cname,_) in enumerate(COLS,1):
            v = row_data.get(cname,"")
            v = "" if pd.isna(v) else str(v)
            val(ws.cell(er,ci),v,row_bg,bold=(cname=="Tracking Status"))
        ws.row_dimensions[er].height = 16

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

# ── Sheet 4: Inscope ───────────────────────────────────────────
def sheet_inscope(ws, df):
    ws.title = "Inscope "
    ws.sheet_view.showGridLines = False

    COLS = [
        ("Carrier Name",30),("Bill of Lading",16),("Order Number",16),
        ("Tracked",10),("Connection Type",16),("Tracking Method",22),
        ("Active Equipment ID",22),("Pickup City State",22),
        ("Final Destination City State",22),("Tracking Status",30),
        ("Milestones Completeness",18),("Milestones Reached",18),
    ]
    STATUS_BG = {"Fully Tracked":"C6EFCE","Partially Tracked":"FFEB9C"}

    for ci,(cname,width) in enumerate(COLS,1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        hdr(ws.cell(1,ci),cname,MID_BLUE,WHITE,True,10,wrap=True)
    ws.row_dimensions[1].height = 25

    for r,(_, row_data) in enumerate(df.iterrows()):
        er = r+2
        status  = str(row_data.get("Tracking Status",""))
        tracked = str(row_data.get("Tracked","")).upper().strip()
        row_bg  = STATUS_BG.get(status,"FCE4D6") if tracked!="TRUE" else STATUS_BG.get(status,WHITE)
        for ci,(cname,_) in enumerate(COLS,1):
            v = row_data.get(cname,"")
            v = "" if pd.isna(v) else str(v)
            val(ws.cell(er,ci),v,row_bg)
        ws.row_dimensions[er].height = 16

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

# ── Build workbook ─────────────────────────────────────────────
def build_workbook(df_raw):
    df = enrich(df_raw)
    wb = Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Pivot for Tracker")
    ws2 = wb.create_sheet("Summary")
    ws3 = wb.create_sheet("FTL data Quality")
    ws4 = wb.create_sheet("Inscope ")
    sheet_pivot(ws1,df)
    sheet_summary(ws2,df)
    sheet_ftl(ws3,df)
    sheet_inscope(ws4,df)
    out = BytesIO(); wb.save(out); out.seek(0)
    return out, df

# ── UI ─────────────────────────────────────────────────────────
uploaded = st.file_uploader("Upload Raw File (Data Quality by Carrier)", type=["xlsx","csv"])

if uploaded:
    with st.spinner("Reading file..."):
        try:
            if uploaded.name.endswith(".csv"):
                df_raw = pd.read_csv(uploaded, header=0)
            else:
                df_raw = pd.read_excel(uploaded, sheet_name=0, header=0)
            df_raw.columns = [c.strip() for c in df_raw.columns]
        except Exception as e:
            st.error(f"Could not read file: {e}"); st.stop()

    with st.expander("📋 Columns detected in uploaded file"):
        st.write(list(df_raw.columns))

    total     = len(df_raw)
    t_bool    = df_raw["Tracked"].astype(str).str.upper().str.strip() == "TRUE"
    tracked   = int(t_bool.sum())
    not_t     = total - tracked
    pct       = round(tracked/total*100,1) if total>0 else 0

    st.success(f"✅ **{total} shipments** loaded")
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Total Shipments",total)
    c2.metric("Tracked",tracked)
    c3.metric("Not Tracked",not_t)
    c4.metric("Tracking Rate",f"{pct}%")

    st.subheader("Carrier Summary Preview")
    cs = (df_raw.groupby("Carrier Name")
                .apply(lambda g: pd.Series({
                    "Total": len(g),
                    "Tracked": (g["Tracked"].astype(str).str.upper()=="TRUE").sum(),
                    "Not Tracked": (g["Tracked"].astype(str).str.upper()!="TRUE").sum(),
                })).reset_index())
    cs["Tracking %"] = (cs["Tracked"]/cs["Total"]*100).round(1).astype(str)+"%"
    cs = cs.sort_values("Total",ascending=False)
    st.dataframe(cs, use_container_width=True, hide_index=True)

    st.subheader("Generate Enriched File")

    if st.button("🚀 Generate Enriched Excel", type="primary"):
        with st.spinner("Building enriched report — please wait..."):
            out_bytes, df_e = build_workbook(df_raw)

        fname = "Unilever_DQ_file.xlsx"
        st.download_button(
            label=f"⬇️ Download {fname}",
            data=out_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success(f"✅ **{fname}** is ready!")

        st.subheader("RCA Breakdown (Not Tracked shipments)")
        rca = (df_e[df_e["Tracked_bool"]==False]["Tracking Status"]
                .value_counts().reset_index())
        rca.columns = ["Root Cause","Count"]
        rca["% of Not Tracked"] = (rca["Count"]/not_t*100).round(1).astype(str)+"%"
        st.dataframe(rca, use_container_width=True, hide_index=True)

else:
    st.info("👆 Upload the raw DQ file above to get started.")
    st.markdown("""
**What this tool produces — 4 sheets matching your weekly file:**

| Sheet | Description |
|---|---|
| **Pivot for Tracker** | Carrier-level tracked vs not tracked + full RCA status breakdown |
| **Summary** | Overall % tracked / not tracked with every RCA category |
| **FTL data Quality** | Full shipment-level detail with milestones, tracking status & errors |
| **Inscope** | Condensed in-scope shipment view |
    """)
